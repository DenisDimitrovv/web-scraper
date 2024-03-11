from bs4 import BeautifulSoup
import requests
import openpyxl

page_to_scrape = requests.get("https://www.imoti.net/bg/obiavi/r/prodava/burgas/?sid=gBkI5L")
soup = BeautifulSoup(page_to_scrape.text, 'html.parser')

offer_types = soup.findAll("span", attrs={"class":"re-offer-type"})
locations = soup.findAll("span", attrs={"class":"location"})
prices = soup.findAll("strong", attrs={"class":"price"})
parameters = soup.findAll("ul", attrs={"class":"parameters"})

workbook_name = "result.xlsx"

wb = openpyxl.load_workbook(workbook_name)

sheet = wb["Results"]

sheet["A1"] = "Offer Types"
sheet["B1"] = "Locations"
sheet["C1"] = "Prices"

for i in range(0, len(prices)):
    offer_type = offer_types[i].text
    location = locations[i].text
    price = prices[i].text.strip()
    sheet.cell(row=i+2, column=1).value = offer_type
    sheet.cell(row=i+2, column=2).value = location
    sheet.cell(row=i+2, column=3).value = price

wb.save(workbook_name)

print("Scraping done")
print('neshto noveo')
    